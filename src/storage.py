"""
storage.py — Excel-backed persistence layer.

Responsibilities
----------------
  1. Create a new workbook with a styled header row if none exists.
  2. Append new job records without overwriting existing data.
  3. Deduplicate on Job URL before writing.
  4. Apply per-column formatting (widths, wrap-text, dropdowns, date format).
  5. Load existing records into memory for in-process deduplication.

Dependencies: pandas, openpyxl
"""

import os
import threading
from datetime import date, datetime
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from config.settings import (
    COLUMNS, COLUMN_WIDTHS, OUTPUT_FILE, SHEET_NAME, STATUS_OPTIONS,
)
from src.utils import get_logger

log = get_logger(__name__)

# ─── Colour palette for the header row ────────────────────────────────────────
HEADER_BG    = "1F3864"   # dark navy
HEADER_FONT  = "FFFFFF"   # white text
ALT_ROW_BG   = "EEF2F7"   # very light blue for even rows
WORKBOOK_LOCK = threading.RLock()


class WorkbookLockedError(PermissionError):
    """Raised when the Excel workbook is locked by another program."""


def _locked_message(filepath: str) -> str:
    return (
        f"Excel workbook is locked: {filepath}. "
        "Close the file in Excel and try again."
    )


# ─── Column index helpers ─────────────────────────────────────────────────────

def _col_idx(col_name: str) -> int:
    """Return 1-based column index for a column name."""
    return COLUMNS.index(col_name) + 1   # COLUMNS is 0-based list


# ─── Workbook helpers ─────────────────────────────────────────────────────────

def _apply_header_style(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Style the header row with bold text, background colour, and borders."""
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FONT, size=11, name="Calibri")
    thin_side   = Side(style="thin", color="AAAAAA")
    thin_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side, bottom=thin_side,
    )
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = col_name
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=False)

    ws.row_dimensions[1].height = 22


def _apply_column_widths(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Set column widths from COLUMN_WIDTHS config."""
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(col_name, 16)


def _add_status_validation(ws: openpyxl.worksheet.worksheet.Worksheet,
                            max_row: int = 5000) -> None:
    """
    Add a dropdown data-validation to the Status column so users can pick
    from the predefined list directly inside Excel.
    """
    status_col = _col_idx("Status")
    status_letter = get_column_letter(status_col)

    # Excel formula-style list: "Option1,Option2,…"
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,  # False = show the arrow button
    )
    dv.add(f"{status_letter}2:{status_letter}{max_row}")
    ws.add_data_validation(dv)


def _freeze_header(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Freeze the header row and first column so they stay visible."""
    ws.freeze_panes = "B2"


def _create_workbook(filepath: str) -> openpyxl.Workbook:
    """Create a fresh, formatted workbook and return it (not yet saved)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _apply_header_style(ws)
    _apply_column_widths(ws)
    _add_status_validation(ws)
    _freeze_header(ws)

    log.info("Created new workbook at %s", filepath)
    return wb


def _open_or_create_workbook(filepath: str):
    """
    Return (wb, ws) — open existing file or create a new one.
    Ensures the target sheet exists.
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _apply_header_style(ws)
            _apply_column_widths(ws)
            _add_status_validation(ws)
            _freeze_header(ws)
        else:
            ws = wb[SHEET_NAME]
    else:
        wb = _create_workbook(filepath)
        ws = wb[SHEET_NAME]

    return wb, ws


def _existing_urls(ws: openpyxl.worksheet.worksheet.Worksheet) -> set:
    """
    Read all Job URLs already stored in the sheet.
    Used to skip duplicates before writing.
    """
    job_url_col = _col_idx("Job URL")
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[job_url_col - 1]   # iter_rows is 0-based within the tuple
        if val:
            urls.add(str(val).strip())
    return urls


def _next_id(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Return the next auto-increment ID value."""
    max_row = ws.max_row
    if max_row < 2:
        return 1
    id_col = _col_idx("ID")
    last_id = ws.cell(row=max_row, column=id_col).value
    if isinstance(last_id, int):
        return last_id + 1
    # Scan backwards to find the last numeric ID
    for row in range(max_row, 1, -1):
        val = ws.cell(row=row, column=id_col).value
        if isinstance(val, int):
            return val + 1
    return 1


def _row_fill(row_num: int) -> Optional[PatternFill]:
    """Alternate row background for readability."""
    if row_num % 2 == 0:
        return PatternFill("solid", fgColor=ALT_ROW_BG)
    return None


def _write_row(ws, row_num: int, record: dict) -> None:
    """
    Write a single job record dict into the worksheet at row_num.
    Applies text wrapping to long columns and alternating row fill.
    """
    thin_side   = Side(style="thin", color="DDDDDD")
    thin_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side, bottom=thin_side,
    )
    fill = _row_fill(row_num)

    # Map dict keys → column names
    col_map = {
        "ID":               record.get("_id", ""),
        "Recruiter Name":   record.get("recruiter_name", ""),
        "Company":          record.get("company", ""),
        "Role":             record.get("role", ""),
        "Stipend/Salary":   record.get("stipend", ""),
        "Location":         record.get("location", ""),
        "Job Type":         record.get("job_type", ""),
        "Email":            record.get("email", ""),
        "LinkedIn URL":     record.get("linkedin_url", ""),
        "Job URL":          record.get("job_url", ""),
        "Description":      record.get("description", ""),
        "Requirements":     record.get("requirements", ""),
        "Date Posted":      record.get("date_posted", ""),
        "Date Scraped":     record.get("date_scraped", date.today().isoformat()),
        "Cold Email Sent":  record.get("cold_email_sent", False),
        "Email Date":       record.get("email_date", ""),
        "Follow-up Sent":   record.get("followup_sent", False),
        "Follow-up Date":   record.get("followup_date", ""),
        "Notes":            record.get("notes", ""),
        "Status":           record.get("status", "Not Contacted"),
    }

    WRAP_COLS = {"Description", "Requirements", "Notes", "Job URL", "LinkedIn URL"}

    for col_name, value in col_map.items():
        col_idx = COLUMNS.index(col_name) + 1
        cell    = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border    = thin_border
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=(col_name in WRAP_COLS),
        )
        if fill:
            cell.fill = fill

    # Set row height taller for wrapped rows
    ws.row_dimensions[row_num].height = 40


# ─── Public API ───────────────────────────────────────────────────────────────

class ExcelStorage:
    """
    Append-safe Excel storage for scraped job records.

    Usage::

        store = ExcelStorage()
        store.save(jobs_list)          # jobs_list from scraper
        df = store.load_dataframe()    # load everything back as DataFrame
    """

    def __init__(self, filepath: str = OUTPUT_FILE):
        self.filepath = filepath

    def save(self, jobs: list[dict]) -> int:
        """
        Append new jobs to the Excel file.
        Deduplicates by Job URL — existing entries are never overwritten.

        Returns the number of new rows written.
        """
        if not jobs:
            log.info("No jobs to save.")
            return 0

        with WORKBOOK_LOCK:
            wb, ws = _open_or_create_workbook(self.filepath)
            existing = _existing_urls(ws)

            new_count  = 0
            skip_count = 0

            for job in jobs:
                url = job.get("job_url", "").strip()
                if url and url in existing:
                    skip_count += 1
                    continue

                row_num   = ws.max_row + 1
                job["_id"] = _next_id(ws)
                _write_row(ws, row_num, job)

                if url:
                    existing.add(url)
                new_count += 1

            try:
                wb.save(self.filepath)
            except PermissionError as exc:
                raise WorkbookLockedError(_locked_message(self.filepath)) from exc
        log.info(
            "Saved %d new rows to %s  (skipped %d duplicates).",
            new_count, self.filepath, skip_count,
        )
        return new_count

    def load_dataframe(self) -> pd.DataFrame:
        """
        Load the entire sheet into a pandas DataFrame.
        Returns an empty DataFrame if the file doesn't exist yet.
        """
        if not os.path.exists(self.filepath):
            log.warning("File not found: %s", self.filepath)
            return pd.DataFrame(columns=COLUMNS)

        try:
            with WORKBOOK_LOCK:
                df = pd.read_excel(self.filepath, sheet_name=SHEET_NAME, dtype=str)
        except PermissionError as exc:
            raise WorkbookLockedError(_locked_message(self.filepath)) from exc
        df = df.fillna("")
        return df

    def update_row(self, job_id: int, updates: dict) -> bool:
        """
        Update specific fields on an existing row identified by ID.
        Useful for the outreach tracker to mark emails sent, etc.

        Args:
            job_id:  The integer ID in the ID column.
            updates: Dict of {column_name: new_value}.

        Returns True on success, False if ID not found.
        """
        if not os.path.exists(self.filepath):
            log.error("File not found: %s", self.filepath)
            return False

        with WORKBOOK_LOCK:
            wb, ws = _open_or_create_workbook(self.filepath)
            id_col = _col_idx("ID")

            for row_num in range(2, ws.max_row + 1):
                cell_id = ws.cell(row=row_num, column=id_col).value
                if cell_id == job_id:
                    for col_name, value in updates.items():
                        if col_name in COLUMNS:
                            col_idx = _col_idx(col_name)
                            ws.cell(row=row_num, column=col_idx).value = value
                    try:
                        wb.save(self.filepath)
                    except PermissionError as exc:
                        raise WorkbookLockedError(_locked_message(self.filepath)) from exc
                    log.info("Updated row ID=%d: %s", job_id, updates)
                    return True

        log.warning("ID=%d not found in %s", job_id, self.filepath)
        return False

    def filter_internships(self) -> pd.DataFrame:
        """Return only rows where Job Type contains 'Internship'."""
        df = self.load_dataframe()
        mask = df["Job Type"].str.contains("internship", case=False, na=False)
        return df[mask].copy()

    def search(self, keyword: str) -> pd.DataFrame:
        """
        Full-text search across Role, Company, Description, and Requirements.
        Returns matching rows as a DataFrame.
        """
        df  = self.load_dataframe()
        kw  = keyword.lower()
        cols = ["Role", "Company", "Description", "Requirements", "Notes"]
        mask = pd.Series(False, index=df.index)
        for col in cols:
            if col in df.columns:
                mask |= df[col].str.lower().str.contains(kw, na=False)
        return df[mask].copy()
